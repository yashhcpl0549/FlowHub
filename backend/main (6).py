#!/usr/bin/env python3
"""
Sale Register Agent - Main Processing Script
Honasa Flow Hub

User uploads (identified by filename patterns):
  - KE30 files       : any file with 'ke30' in name (1 to N files)
  - Customer Mapping : file with 'customer_mapping' or 'final_customer' in name
  - MRP file         : file with 'zmrp' or 'mrp' in name
  - Freebie file     : file with 'freebie' or 'bxgy' in name

Auto-fetched (no upload needed):
  - GST Mapping    : Google Sheet "Fountain 9 Master", tab "Product Master"
  - Product Master : BigQuery `mamaearth-262312.SAP.Product_Master`
  - Template       : gs://honasa-flow-hub/ke30_template/Sale_Register_Template.xlsx

Output:
  - Sale_Register_Filled_{mon}{yy}.xlsx
"""

import sys, json, os, shutil, zipfile, tempfile, calendar, logging, warnings, datetime
import numpy as np
import pandas as pd
from pathlib import Path
import xml.etree.ElementTree as ET
from openpyxl import Workbook
import io
from datetime import date
from google.cloud import storage, bigquery
from google.oauth2.service_account import Credentials

warnings.filterwarnings("ignore")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger(__name__)

# ── Constants ─────────────────────────────────────────────────────────────────
GCS_BUCKET        = "honasa-flow-hub"
TEMPLATE_GCS_PATH = "ke30_template/Sale_Register_Template.xlsx"
BQ_PROJECT        = "mamaearth-262312"
BQ_PM_TABLE       = "mamaearth-262312.SAP.Product_Master"
BQ_GST_TABLE      = "mamaearth-262312.reporting_tables.BI_F9_Product_Master_Import"
CREDENTIALS_PATH  = os.environ.get("GCS_CREDENTIALS_PATH", "/app/credentials/gcs-credentials.json")

# ── File helpers ──────────────────────────────────────────────────────────────
def find_file(files, *patterns):
    """Returns local_path of first matching file (set after GCS download)."""
    for f in files:
        name = f["filename"].lower()
        for p in patterns:
            if p.lower() in name:
                return f.get("local_path") or ""
    return None

def find_files(files, *patterns):
    """Returns list of local_paths of all matching files (set after GCS download)."""
    result = []
    for f in files:
        name = f["filename"].lower()
        for p in patterns:
            if p.lower() in name:
                result.append(f.get("local_path") or "")
                break
    return result

def file_exists_by_pattern(files, *patterns):
    """Check if any file matches pattern — works before GCS download."""
    for f in files:
        name = f["filename"].lower()
        for p in patterns:
            if p.lower() in name:
                return True
    return False

# ── Step 00: Download input files from GCS to local temp dir ──────────────────
def download_inputs_from_gcs(files, dest_dir):
    log.info(f"Downloading {len(files)} input file(s) from GCS...")
    client = storage.Client.from_service_account_json(CREDENTIALS_PATH)
    bucket = client.bucket(GCS_BUCKET)
    for f in files:
        gcs_path   = f.get("gcs_path")
        filename   = f["filename"]
        local_path = os.path.join(dest_dir, filename)
        if gcs_path:
            blob = bucket.blob(gcs_path)
            blob.download_to_filename(local_path)
            f["local_path"] = local_path
            log.info(f"  ✅ {filename}")
        else:
            # Already has a local_path (e.g. during direct testing)
            log.info(f"  ↩  {filename} (using existing local path)")
    return files

def read_excel_smart(path, **kwargs):
    df_test = pd.read_excel(path, engine="openpyxl", nrows=1)
    if df_test.iloc[0].isna().all() or str(df_test.columns[0]).startswith("Unnamed"):
        return pd.read_excel(path, engine="openpyxl", skiprows=1, **kwargs)
    return pd.read_excel(path, engine="openpyxl", **kwargs)

# ── Step 0a: Download template from GCS ───────────────────────────────────────
def download_template(dest_dir):
    log.info(f"Downloading template from GCS: gs://{GCS_BUCKET}/{TEMPLATE_GCS_PATH}")
    client = storage.Client.from_service_account_json(CREDENTIALS_PATH)
    blob   = client.bucket(GCS_BUCKET).blob(TEMPLATE_GCS_PATH)
    local  = os.path.join(dest_dir, "Sale_Register_Template.xlsx")
    blob.download_to_filename(local)
    log.info("  ✅ Template downloaded")
    return local

# ── Step 0b: Fetch GST mapping from BigQuery ──────────────────────────────────
def fetch_gst_mapping():
    log.info(f"Fetching GST mapping from BigQuery: {BQ_GST_TABLE}")
    bq_client = bigquery.Client.from_service_account_json(CREDENTIALS_PATH, project=BQ_PROJECT)
    query = f"SELECT material_code, gst_code FROM `{BQ_GST_TABLE}` WHERE material_code IS NOT NULL"
    gst = bq_client.query(query).to_dataframe()
    log.info(f"  ✅ {len(gst)} rows fetched from BigQuery")
    gst = gst[["material_code", "gst_code"]].drop_duplicates()
    gst["material_code"] = pd.to_numeric(gst["material_code"], errors="coerce").astype("Int64")
    gst = gst[gst["material_code"].notna()]
    return gst

# ── Step 0c: Fetch Product Master from BigQuery ───────────────────────────────
def fetch_product_master():
    log.info(f"Fetching Product Master from BigQuery: {BQ_PM_TABLE}")
    bq_client = bigquery.Client.from_service_account_json(CREDENTIALS_PATH, project=BQ_PROJECT)
    query = f"""
        SELECT Material, ean_code, product_name, brand, mrp,
               material_group, price, category, sub_category
        FROM `{BQ_PM_TABLE}`
    """
    PM = bq_client.query(query).to_dataframe()
    log.info(f"  ✅ Product Master: {len(PM)} rows fetched from BigQuery")
    return PM

# ── Step 1: Load & combine KE30 files ─────────────────────────────────────────
def load_ke30_files(ke30_paths):
    log.info(f"Loading {len(ke30_paths)} KE30 file(s)...")
    frames = []
    for path in ke30_paths:
        try:
            df = read_excel_smart(path)
            df["source_file"] = os.path.basename(path)
            log.info(f"  ✅ {os.path.basename(path)}: {df.shape}")
            frames.append(df)
        except Exception as e:
            log.warning(f"  ⚠️  Skipping {os.path.basename(path)}: {e}")
    if not frames:
        raise ValueError("No KE30 files could be loaded.")
    ke30 = pd.concat(frames, ignore_index=True)
    if "Sales order item" not in ke30.columns and "Sales Order Item" in ke30.columns:
        ke30.rename(columns={"Sales Order Item": "Sales order item"}, inplace=True)
    log.info(f"  Combined KE30: {ke30.shape}")
    return ke30

# ── Step 2: Customer mapping ──────────────────────────────────────────────────
def merge_customer_mapping(ke30, path):
    log.info("Merging customer mapping...")
    cm = pd.read_csv(path)[["Customer Code", "Channel", "Updated Strat Heads For MIS"]].drop_duplicates()
    ke30["Customer"] = ke30["Customer"].astype(str).str.replace(r"\.0+$", "", regex=True)
    ke30.loc[ke30["Customer"] == "nan", "Customer"] = "1103927"
    cm["Customer Code"] = cm["Customer Code"].astype(str).str.replace(r"\.0+$", "", regex=True)
    merged = pd.merge(ke30, cm, how="left", left_on="Customer", right_on="Customer Code")
    missing = merged["Channel"].isna().sum()
    if missing:
        log.warning(f"  ⚠️  {missing} rows without channel mapping")
    return merged

# ── Step 3: GST mapping ───────────────────────────────────────────────────────
def merge_gst_mapping(df, gst_df):
    log.info("Merging GST mapping...")
    df["Product"] = pd.to_numeric(df["Product"], errors="coerce").astype("Int64")
    merged = pd.merge(df, gst_df, how="left", left_on="Product", right_on="material_code")
    merged.drop(columns="material_code", inplace=True, errors="ignore")
    merged["gst_code"] = merged["gst_code"].replace("", 18).fillna(18)
    return merged

# ── Step 4: MRP merge ─────────────────────────────────────────────────────────
def _date_range(ke30):
    posting     = pd.to_datetime(ke30["Posting date"], errors="coerce").dropna()
    min_d       = posting.min().date()
    max_d       = posting.max().date()
    frame_start = min_d.replace(day=1)
    frame_end   = max_d.replace(day=calendar.monthrange(max_d.year, max_d.month)[1])
    log.info(f"  Date range: {frame_start} → {frame_end}")
    return frame_start, frame_end

def merge_mrp(ke30_gst, mrp_path):
    log.info("Processing MRP...")
    mrp = pd.read_excel(mrp_path, engine="openpyxl")
    mrp["Validity To"]   = mrp["Validity To"].astype(str).str.slice(0, 10)
    mrp["Validity From"] = mrp["Validity From"].astype(str).str.slice(0, 10)
    mrp.loc[mrp["Validity To"] > str(date.today()), "Validity To"] = str(date.today())
    mrp = mrp[mrp["Validity From"] != "9999-12-31"]
    mrp["Validity To"]   = pd.to_datetime(mrp["Validity To"]).dt.date
    mrp["Validity From"] = pd.to_datetime(mrp["Validity From"]).dt.date
    mrp["Material Number"] = mrp["Material Number"].astype(str).str.lstrip("0")

    frame_start, frame_end = _date_range(ke30_gst)

    mrp_f = mrp[
        mrp["Plant"].notna() & (mrp["Plant"] != "") & (mrp["Condition type"] == "ZMRP") &
        (mrp["Validity From"] <= frame_end) &
        (mrp["Validity To"].isna() | (mrp["Validity To"] >= frame_start))
    ].sort_values("Validity From", ascending=False).sort_values("Plant")
    mrp_f.drop(columns=["Control code", "Condition type"], inplace=True, errors="ignore")
    mrp_f = mrp_f.drop_duplicates()
    mrp_f.rename(columns={"Validity To": "valid_to", "Validity From": "validity_from"}, inplace=True)

    mrp_f["Cap_validity_from"] = mrp_f["validity_from"].apply(lambda x: frame_start if x <= frame_start else x)
    mrp_f["Cap_validity_to"]   = mrp_f["valid_to"].apply(lambda x: frame_end if x >= frame_end else x)
    mrp_f = mrp_f.drop(columns=["valid_to", "validity_from"]).drop_duplicates()
    mrp_f.rename(columns={"Cap_validity_to": "valid_to", "Cap_validity_from": "validity_from"}, inplace=True)

    mrp_f["Date"] = mrp_f.apply(lambda r: pd.date_range(start=r["validity_from"], end=r["valid_to"]), axis=1)
    mrp_exp = mrp_f.explode("Date").reset_index(drop=True)
    mrp_exp.drop(columns=["validity_from", "valid_to"], inplace=True)
    mrp_exp.rename(columns={"Material Number": "MaterialNumber"}, inplace=True)
    mrp_exp[["Plant", "MaterialNumber"]] = mrp_exp[["Plant", "MaterialNumber"]].astype(str)

    ke30_gst["Plant"]   = ke30_gst["Plant"].astype(str).str.replace(r"\.0+$", "", regex=True)
    ke30_gst["Product"] = ke30_gst["Product"].astype(str).str.replace(r"\.0+$", "", regex=True)

    result = pd.merge(ke30_gst, mrp_exp, how="left",
                      left_on=["Plant", "Product", "Posting date"],
                      right_on=["Plant", "MaterialNumber", "Date"])
    result.drop(columns=["Customer Code", "MaterialNumber", "Date"], inplace=True, errors="ignore")
    log.info(f"  KE30 after MRP merge: {result.shape}")
    return result

# ── Step 5: Sample tagging ────────────────────────────────────────────────────
def tag_samples(ke30_final):
    log.info("Tagging samples...")
    billing_list = ["ZFOC", "ZFRC", "ZRFC", "ZUFC"]
    df_in  = ke30_final[ke30_final["Billing Type"].isin(billing_list)].copy()
    df_out = ke30_final[~ke30_final["Billing Type"].isin(billing_list)].copy()
    df_in["Sample"] = "Yes"

    df_out1 = df_out.copy()
    df_out2 = df_out[df_out["Distribution Channel"].astype(str).isin(["30"])].copy()
    df_out1["Abs qty"] = df_out1["Total Quantity"].abs()

    mask_pos = (
        (df_out1["Total Quantity"].abs() > 0) &
        (df_out1["Val/COArea Crcy"].abs() / df_out1["Total Quantity"].abs() < 10) &
        (df_out1["Reference procedure"] == "VBRK")
    )
    df_pos = df_out1[mask_pos].copy()
    df_neg = df_out1[~mask_pos].copy()
    df_pos["Select"] = "1"

    keys = ["Sales Order", "Sales order item", "Posting date", "Abs qty", "Billing Type"]
    df_merged = df_neg.merge(df_pos[keys + ["Select"]], on=keys, how="left")

    for df in [df_merged, df_neg, df_pos]:
        df.drop(columns=["Abs qty"], inplace=True, errors="ignore")

    mkpf_sel = (df_merged["Reference procedure"] == "MKPF") & (df_merged["Select"] == "1")
    df_mkpf  = pd.concat([df_merged[mkpf_sel], df_pos], ignore_index=True)
    df_na    = df_merged[~mkpf_sel].copy()
    df_na["Sample"]   = "No"
    df_mkpf["Sample"] = "Yes"
    for df in [df_na, df_mkpf, df_merged]:
        df.drop(columns=["Select"], inplace=True, errors="ignore")

    return df_in, df_out2, df_na, df_mkpf

# ── Step 6: D2C Freebie processing ───────────────────────────────────────────
def process_freebies(freebie_path, PM, df_out2, ke30_original):
    log.info("Processing D2C freebies...")
    freebie = pd.concat(pd.read_excel(freebie_path, sheet_name=None).values(), ignore_index=True)
    freebie["quantity"] = freebie["quantity"].astype(float)
    freebie_backup = freebie.copy()

    freebie.drop(columns=["MRP"], inplace=True)
    freebie = freebie.groupby(["year_month", "coupon_code", "SKU", "brand"], as_index=False).agg({"quantity": "sum"})

    PM = PM[["Material", "ean_code", "product_name", "brand",
               "mrp", "material_group", "price", "category", "sub_category"]].copy()
    gst_ean = PM[["ean_code", "Material"]].drop_duplicates(subset=["ean_code"]).rename(columns={"Material": "material_code"})

    freebie["SKU"] = freebie["SKU"].astype(str)
    freebie = freebie.merge(gst_ean, how="left", left_on="SKU", right_on="ean_code")
    freebie.rename(columns={"material_code": "Product"}, inplace=True)

    df_d2c_temp = df_out2[df_out2["Distribution Channel"].astype(str).isin(["30"])].copy()
    df_d2c_temp["mkpf_abs_qty"] = np.where(
        df_d2c_temp["Reference procedure"] == "MKPF",
        df_d2c_temp["Total Quantity"].abs(), 0
    )
    df_agg = (df_d2c_temp.groupby("Product", as_index=False)["mkpf_abs_qty"].sum()
              .rename(columns={"mkpf_abs_qty": "MKPF_Abs_Qty_Sum"}))

    freebie = freebie.merge(df_agg, on="Product", how="left")
    freebie["Total Quantity"] = freebie[["quantity", "MKPF_Abs_Qty_Sum"]].min(axis=1)
    freebie.drop(columns=["quantity", "MKPF_Abs_Qty_Sum"], inplace=True)

    freebie["year_month"] = pd.to_datetime(freebie["year_month"].astype(str) + "-01")
    freebie.rename(columns={"year_month": "Posting date"}, inplace=True)
    freebie["Segment Text"] = freebie["Division Text"] = freebie["brand"]

    freebie_backup["SKU"] = freebie_backup["SKU"].astype(str)
    sku_mrp = (freebie_backup.sort_values(["SKU", "quantity"], ascending=[True, False])
               .drop_duplicates(subset=["SKU"])[["SKU", "MRP"]].reset_index(drop=True))
    freebie = freebie.merge(sku_mrp, on="SKU", how="left")
    freebie.drop(columns=["coupon_code", "SKU", "brand", "ean_code"], inplace=True, errors="ignore")
    freebie = freebie[freebie["Product"].notna()]

    df_d2c = df_out2[df_out2["Distribution Channel"].astype(str).isin(["30"])].copy()
    n = len(freebie)
    row_vbrk = df_d2c[df_d2c["Reference procedure"] == "VBRK"].sample(1, random_state=42).iloc[0]
    row_mkpf = df_d2c[df_d2c["Reference procedure"] == "MKPF"].sample(1, random_state=42).iloc[0]

    row_map   = pd.DataFrame([row_vbrk, row_mkpf])
    tiled     = pd.DataFrame(np.tile(row_map.values, (n, 1)), columns=row_map.columns)
    rep       = pd.DataFrame(np.repeat(freebie.values, 2, axis=0), columns=freebie.columns)

    ff = pd.DataFrame(columns=df_d2c.columns)
    common = freebie.columns.intersection(df_d2c.columns)
    ff[common] = rep[common].reset_index(drop=True)
    for col in [c for c in df_d2c.columns if c not in common]:
        ff[col] = tiled[col].values if col in tiled.columns else pd.NA
    ff = ff.reset_index(drop=True)

    pm_cols = ["Material", "ean_code", "product_name", "brand", "mrp", "material_group", "price", "category", "sub_category"]

    ff["Product"] = ff["Product"].astype(str)
    ff = ff.merge(PM, left_on="Product", right_on="Material", how="left")
    ff["Customer Text"] = ff["brand"] + " D2C"
    ff["Brand Text"]    = ff["brand"]
    ff["Material Group"] = ff["material_group"]
    ff["Material Group Text"] = ff["category"] + " - " + ff["sub_category"]
    ff["Product Text"]  = ff["product_name"]
    ff["Segment Text"]  = ff["Division Text"] = ff["brand"]
    ff.drop(columns=pm_cols, inplace=True, errors="ignore")

    ff.loc[ff["Reference procedure"] == "VBRK", ["Val/COArea Crcy", "Amount"]] = 0

    ke30_o = ke30_original.copy()
    if "Sales order item" not in ke30_o.columns and "Sales Order Item" in ke30_o.columns:
        ke30_o.rename(columns={"Sales Order Item": "Sales order item"}, inplace=True)
    g = (ke30_o[ke30_o["Reference procedure"] == "MKPF"]
         .groupby("Product", as_index=False)
         .agg({"Val/COArea Crcy": "sum", "Total Quantity": "sum"}))
    g["COGS"] = g["Val/COArea Crcy"] / g["Total Quantity"]
    ff = ff.merge(g[["Product", "COGS"]], on="Product", how="left")
    ff.loc[ff["Reference procedure"] == "MKPF", "Val/COArea Crcy"] = ff.loc[ff["Reference procedure"] == "MKPF", "COGS"]
    ff.drop(columns=["COGS"], inplace=True)

    ff = ff.merge(PM, left_on="Product", right_on="Material", how="left")
    ff = ff[~((ff["Reference procedure"] == "MKPF") & ff["Val/COArea Crcy"].isna())].copy()

    ff["Val/COArea Crcy"] = ff["Val/COArea Crcy"].astype(str).str.split(".").str[0].astype(float)
    ff["Val/COArea Crcy"] *= ff["Total Quantity"]
    ff.loc[ff["Reference procedure"] == "VBRK", "Val/COArea Crcy"] = 0
    ff.loc[ff["Amount"] == 0, "Amount"] = ff["mrp"]
    ff.drop(columns=pm_cols, inplace=True, errors="ignore")

    ff.loc[ff["Reference procedure"] == "VBRK", "Val/COArea Crcy"] *= -1
    ff.loc[ff["Reference procedure"] == "VBRK", "Total Quantity"]  *= -1

    brand_df = pd.DataFrame({
        "Brand Text": ["Mamaearth","Staze","The Derma Co.","Aqualogica","Dr. Sheth's","BBLUNT","Pure Origin","Ayuga","Mamaearth Fresh","Lumineve"],
        "Segment":   ["ME","SZ","DE","AQ","DS","BB","PO","AY","MF","LM"],
        "Division":  [10,70,20,30,21,50,80,40,10,90],
    })

    ff_copy = ff.copy()
    ff_copy["Val/COArea Crcy"] *= -1
    ff_copy["Total Quantity"]  *= -1
    ff_copy["Sample"]   = "No"
    ff["Sample"]        = "Yes"
    ff["Customer"]      = "B1G1"
    ff_copy["Customer"] = "dummy_B1G1"

    ff_copy.drop(columns=["Segment", "Division"], inplace=True, errors="ignore")
    ff_copy = ff_copy.merge(brand_df, on="Brand Text", how="left")

    result = pd.concat([ff, ff_copy], ignore_index=True)
    log.info(f"  Freebie rows: {len(result)}")
    return result

# ── Step 7: Build mega_df & split sheets ──────────────────────────────────────
def build_mega_df(df_freebie, df_na, df_mkpf, df_in):
    log.info("Building mega dataframe...")
    mega = pd.concat([df_freebie, df_na, df_mkpf, df_in], ignore_index=True)
    mega["Sample"] = mega["Sample"].fillna("No")

    mega["GST"] = np.where(
        mega["Reference procedure"] == "VBRK",
        mega["Val/COArea Crcy"].astype(float) * mega["gst_code"].astype(float) / 100, 0)
    mega["GMV"] = np.where(
        mega["Reference procedure"] == "VBRK",
        mega["Total Quantity"].astype(float) * mega["Amount"].astype(float), 0)

    mega.rename(columns={"Amount": "MRP", "Customer Group": "Customer Group Code",
                          "Updated Strat Heads For MIS": "Customer Group"}, inplace=True)
    mega.loc[mega["Reference procedure"] == "VBRK", "Type"] = "Sales"
    mega.loc[mega["Reference procedure"] == "MKPF", "Type"] = "COGS"
    mega.loc[mega["Reference procedure"] == "MKPF", "MRP"]  = 0

    mask_zero = (
        (mega["Billing Type"].isin(["ZCR", "ZDR"]) | (mega["Cost Element"].astype(str) == "40003000")) &
        (mega["Reference procedure"] == "VBRK")
    )
    mega.loc[mask_zero, ["Total Quantity", "GMV"]] = 0

    combo_Z = mega[mega["Billing Type"].isin(["ZCR","ZDR"])][["Sales Order","Sales order item"]].drop_duplicates()
    S_rows  = mega.merge(combo_Z, on=["Sales Order","Sales order item"])[mega["Billing Type"].isin(["S1","S2"])]
    sep_df  = S_rows.copy()
    mega    = mega.drop(S_rows.index)

    mega.loc[mega["Customer"].isin(["1103708"]), "Sample"] = "Damaged"
    mega.loc[mega["Customer"].isin(["1101152","1104217"]), "Sample"] = "Sample Octavos"
    mega.loc[mega["Product"].isin([f"NV{str(i).zfill(6)}" for i in range(45,73)]), "Sample"] = "Yes"
    mega["Brand Text"] = mega["Brand Text"].replace({"Service Sale":"Mamaearth","Derma Co":"The Derma Co.","Acqualogica":"Aqualogica"})
    mega.loc[mega["Distribution Channel"].astype(str) == "80", ["GST","gst_code"]] = [0,"0"]
    mega.loc[mega["Reference procedure"] == "MKPF", "gst_code"] = "0"

    log.info(f"  Mega DF: {mega.shape}")
    s1 = mega[mega["Channel"].isin(["Website","EBO"])].copy()
    s2 = mega[mega["Channel"] == "B2C ECOM"].copy()
    s3 = mega[~mega["Channel"].isin(["Website","EBO","B2C ECOM"])].copy()
    log.info(f"  Sheets → EBO/Web:{len(s1)}  B2C:{len(s2)}  Remaining:{len(s3)}  S1S2:{len(sep_df)}")
    return {"RAW_Website_EBO": s1, "RAW_b2c": s2, "RAW_Remaining": s3, "RAW_S1S2": sep_df}

# ── Step 8: Write to template ─────────────────────────────────────────────────
def _compact_wb(path, sheet_map):
    wb = Workbook(write_only=True)
    for name, df in sheet_map.items():
        df2 = df.replace([np.inf,-np.inf], None).fillna("")
        ws  = wb.create_sheet(title=name)
        ws.append(list(df2.columns))
        for row in df2.itertuples(index=False, name=None):
            ws.append(list(row))
        log.info(f"  ✔ {name}: {len(df2)} rows × {len(df2.columns)} cols")
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    wb.save(path)

def _map_sheets(dirpath):
    rels = ET.parse(str(Path(dirpath)/"xl"/"_rels"/"workbook.xml.rels")).getroot()
    rid_map = {}
    for r in rels:
        if r.tag.endswith("Relationship"):
            tgt = r.attrib["Target"].lstrip("/").replace("../","").replace("xl/xl/","xl/")
            rid_map[r.attrib["Id"]] = ("xl/"+tgt if not tgt.startswith("xl/") else tgt)
    root = ET.parse(str(Path(dirpath)/"xl"/"workbook.xml")).getroot()
    m = {}
    for sh in root.iter("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet"):
        rid = sh.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
        m[sh.attrib["name"]] = rid_map[rid]
    return m

def _extract_sheetdata_text(xml_path):
    """Extract <sheetData>...</sheetData> as raw string + dimension ref.
    Reads file as text — never builds an XML element tree for large files."""
    import re
    with open(xml_path, "r", encoding="utf-8", errors="replace") as f:
        raw = f.read()
    dim_ref = None
    dim_m = re.search(r'<dimension[^>]*ref="([^"]+)"', raw)
    if dim_m:
        dim_ref = dim_m.group(1)
    sd_m = re.search(r'(<sheetData(?:[^>]*)>.*?</sheetData>)', raw, re.DOTALL)
    if not sd_m:
        sd_m = re.search(r'(<sheetData/>)', raw)
    return (sd_m.group(1) if sd_m else "<sheetData/>"), dim_ref

def _inject_sheetdata_text(tmpl_xml_path, sheetdata_text, dim_ref):
    """Replace sheetData block in template XML using string ops.
    Template has no data rows so it is small and safe to load fully."""
    import re
    with open(tmpl_xml_path, "r", encoding="utf-8", errors="replace") as f:
        tmpl = f.read()
    if dim_ref:
        tmpl = re.sub(r'<dimension[^>]*ref="[^"]*"', f'<dimension ref="{dim_ref}"', tmpl)
    tmpl = re.sub(r'<sheetData(?:[^>]*)>.*?</sheetData>', sheetdata_text, tmpl, flags=re.DOTALL)
    tmpl = re.sub(r'<sheetData/>', sheetdata_text, tmpl)
    with open(tmpl_xml_path, "w", encoding="utf-8") as f:
        f.write(tmpl)

def write_to_template(raw_sheets, template_path, output_path):
    log.info("Injecting data into Excel template...")
    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ET.register_namespace("", ns)
    tpl_dir = tempfile.mkdtemp(prefix="tpl_")
    raw_dir = tempfile.mkdtemp(prefix="raw_")
    try:
        with zipfile.ZipFile(template_path,"r") as z: z.extractall(tpl_dir)
        raw_xlsx = os.path.join(raw_dir,"raw_data.xlsx")
        _compact_wb(raw_xlsx, raw_sheets)
        with zipfile.ZipFile(raw_xlsx,"r") as z: z.extractall(raw_dir)
        tpl_map = _map_sheets(tpl_dir)
        raw_map = _map_sheets(raw_dir)
        for sname in raw_sheets:
            if sname not in tpl_map or sname not in raw_map:
                log.warning(f"  ⚠️  '{sname}' missing — skipped"); continue
            tmpl_xml = Path(tpl_dir)/tpl_map[sname]
            raw_rel  = raw_map[sname].lstrip("/")
            if not raw_rel.startswith("xl/"): raw_rel = "xl/"+raw_rel
            raw_xml  = Path(raw_dir)/raw_rel
            sheetdata_text, dim_ref = _extract_sheetdata_text(str(raw_xml))
            _inject_sheetdata_text(str(tmpl_xml), sheetdata_text, dim_ref)
            log.info(f"  ✔ Updated: {sname}")
        tmp = tempfile.mktemp(prefix="final_", suffix=".xlsx")
        with zipfile.ZipFile(tmp,"w",compression=zipfile.ZIP_DEFLATED) as z:
            for folder,_,files in os.walk(tpl_dir):
                for fn in files:
                    fp = os.path.join(folder,fn)
                    z.write(fp, os.path.relpath(fp,tpl_dir).replace("\\","/"), compress_type=zipfile.ZIP_DEFLATED)
        shutil.move(tmp, output_path)
        log.info(f"✅ Output: {output_path}")
    finally:
        shutil.rmtree(tpl_dir, ignore_errors=True)
        shutil.rmtree(raw_dir, ignore_errors=True)

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python main.py <config.json>")
    with open(sys.argv[1]) as fh:
        config = json.load(fh)

    files       = config["files"]
    output_path = config["output_path"]
    os.makedirs(output_path, exist_ok=True)

    log.info("=== Sale Register Agent — Start ===")

    # Check files exist by name BEFORE downloading
    missing = []
    if not file_exists_by_pattern(files, "ke30"):            missing.append("KE30 file(s) — name must contain 'ke30'")
    if not file_exists_by_pattern(files, "customer_mapping", "final_customer"): missing.append("Customer mapping CSV — name must contain 'customer_mapping'")
    if not file_exists_by_pattern(files, "zmrp", "mrp"):    missing.append("MRP file — name must contain 'zmrp' or 'mrp'")
    if not file_exists_by_pattern(files, "freebie", "bxgy"): missing.append("Freebie file — name must contain 'freebie' or 'bxgy'")
    if missing:
        raise SystemExit("❌ Missing required files:\n" + "\n".join(f"  - {m}" for m in missing))

    log.info(f"KE30 files    : {[f['filename'] for f in files if 'ke30' in f['filename'].lower()]}")
    log.info(f"Cust map      : {next(f['filename'] for f in files if any(p in f['filename'].lower() for p in ['customer_mapping','final_customer']))}")
    log.info(f"MRP           : {next(f['filename'] for f in files if any(p in f['filename'].lower() for p in ['zmrp','mrp']))}")
    log.info(f"Freebie       : {next(f['filename'] for f in files if any(p in f['filename'].lower() for p in ['freebie','bxgy']))}")

    tmp_dir = tempfile.mkdtemp(prefix="sr_agent_")
    # Download input files from GCS to local temp dir
    files   = download_inputs_from_gcs(files, tmp_dir)
    # Resolve local paths after download
    ke30_paths            = find_files(files, "ke30")
    customer_mapping_path = find_file(files, "customer_mapping", "final_customer")
    mrp_path              = find_file(files, "zmrp", "mrp")
    freebie_path          = find_file(files, "freebie", "bxgy")

    template_path = download_template(tmp_dir)
    gst_df        = fetch_gst_mapping()
    PM            = fetch_product_master()

    try:
        ke30         = load_ke30_files(ke30_paths)
        ke30_ch      = merge_customer_mapping(ke30, customer_mapping_path)
        ke30_gst     = merge_gst_mapping(ke30_ch, gst_df)
        ke30_final   = merge_mrp(ke30_gst, mrp_path)
        df_in, df_out2, df_na, df_mkpf = tag_samples(ke30_final)
        df_freebie   = process_freebies(freebie_path, PM, df_out2, ke30)
        raw_sheets   = build_mega_df(df_freebie, df_na, df_mkpf, df_in)

        posting      = pd.to_datetime(ke30["Posting date"], errors="coerce").dropna()
        month_str    = posting.min().strftime("%b%y").lower()
        output_file  = os.path.join(output_path, f"Sale_Register_Filled_{month_str}.xlsx")

        write_to_template(raw_sheets, template_path, output_file)
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

    log.info("=== Sale Register Agent — Done ===")

if __name__ == "__main__":
    main()
